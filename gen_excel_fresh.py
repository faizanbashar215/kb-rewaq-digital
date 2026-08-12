#!/usr/bin/env python3
# Build KB_Rewaq_Fresh_Leads.xlsx from fresh_sites CRM records. 8-col pahle-jaisa format.
import os, json, glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import quote
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from dm_engine import compose_dm, site_url_for

CRM = os.environ.get("CLIENTS_DIR", r"D:\KB Rewaq Clients")
SITES = os.environ.get("AGENCY_DIR", r"D:\digitalfirst-agency") + r"\fresh_sites"
OUT = os.environ.get("AGENCY_DIR", r"D:\digitalfirst-agency") + r"\KB_Rewaq_Fresh_Leads.xlsx"
WA_YOU = "96550703252"

leads = []
for d in sorted(glob.glob(os.path.join(CRM, "*/"))):
    cj = os.path.join(d, "client.json")
    if not os.path.isfile(cj):
        continue
    j = json.load(open(cj, encoding="utf-8"))
    biz = j.get("business", {}); con = j.get("contact", {}); loc = j.get("location", {}); onl = j.get("online", {})
    slug = biz.get("slug", os.path.basename(d.rstrip("/\\")))
    phone = con.get("phone", "") or ""
    name = biz.get("name_en", "")
    area = loc.get("area_en", "")
    site = site_url_for(slug)
    # Per-lead unique human DM from shared engine (English, no em-dash, Jarvis intro, automation, no price)
    dm = compose_dm(name, area, slug, len(leads))
    phone_disp = con.get("phone_disp", "") or ("+" + phone if phone else "+" + WA_YOU)
    leads.append({"slug": slug, "name": name, "name_ar": biz.get("name_ar", ""), "area": area,
                  "phone": phone or WA_YOU, "phone_disp": phone_disp, "site": site, "dm": dm})

leads.sort(key=lambda x: x["name"].lower())

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "KB Rewaq Leads"
# SIMPLE clickable layout (like the original 5-lead sheet boss liked):
# col E = Click to DM (WhatsApp opens with DM+site link typed), col F = View Demo site.
headers = ["#", "Business (EN)", "Business (AR)", "Area",
           "Phone & Website", "Click to DM (WhatsApp)", "Ready DM Message", "Status"]
ws.append(headers)

green = Font(color="006100"); green_fill = PatternFill("solid", fgColor="C6EFCE")
blue = Font(color="0563C1", underline="single"); blue_fill = PatternFill("solid", fgColor="DDEBF7")
hdr_fill = PatternFill("solid", fgColor="1F4E78"); hdr_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")

for i, L in enumerate(leads, 1):
    phone_num = L["phone"] or WA_YOU
    phone_link = f"https://wa.me/{phone_num}?text={quote(L['dm'])}"
    demo_link = L["site"] if L["site"] else ""
    phone_disp = ("+" + phone_num) if not phone_num.startswith("+") else phone_num
    # col5: plain text phone + website (visible on GitHub + Excel, copy-paste)
    contact_cell = f"{phone_disp}  |  {demo_link}" if demo_link else phone_disp
    # col6: clickable "Click to DM" -> WhatsApp opens with DM typed
    dm_cell = f'=HYPERLINK("{phone_link}","Click to DM")'
    row = [i, L["name"], L["name_ar"], L["area"], contact_cell, dm_cell, L["dm"], "Live" if L["site"] else "Pending"]
    ws.append(row)
    r = ws.max_row
    for c in range(1, 9):
        ws.cell(r, c).border = border
        ws.cell(r, c).alignment = wrap
    if L["site"]:
        ws.cell(r, 6).font = blue; ws.cell(r, 6).fill = blue_fill
        ws.cell(r, 8).font = Font(color="006100", bold=True)

for c in range(1, 9):
    ws.cell(1, c).fill = hdr_fill; ws.cell(1, c).font = hdr_font
    ws.cell(1, c).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
widths = [5, 26, 20, 16, 42, 18, 60, 11]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 30

htu = wb.create_sheet("HOW TO USE")
htu["A1"] = "KB REWAQ DIGITAL — FRESH LEAD TRACKER"
htu["A1"].font = Font(bold=True, size=14, color="1F4E78")
lines = ["", "1. Column E (Phone copy) = plain number + wa.me link. Copy-paste or open wa.me manually.",
         "2. Column F (Click to DM) = BLUE link. Open this Excel file in Excel/Google Sheets and CLICK it -> WhatsApp opens with the DM already typed. One click sends.",
         "3. Column G (Demo URL) = plain website link to copy.",
         "4. Column H (Demo click) = BLUE link. Click to open their demo site.",
         "5. Column I = full DM text (human, English, Jarvis intro, no price).",
         "6. Send DM via Column F. When they reply YES, visit salon, discuss price, sign up.",
         "7. Change Status (col J) to: replied / won / paid.", "",
         f"Your WA: +{WA_YOU}   |   Brand: KB Rewaq Digital",
         "", f"Total fresh leads: {len(leads)}   |   Live demo sites: {sum(1 for x in leads if x['site'])}"]
for i, t in enumerate(lines, 2):
    htu.cell(i, 1).value = t
htu.column_dimensions["A"].width = 110

wb.save(OUT)
print(f"SAVED {OUT} | Total {len(leads)} | Live {sum(1 for x in leads if x['site'])}")
