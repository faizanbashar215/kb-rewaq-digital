#!/usr/bin/env python3
# Import OpenCode-agent leads xlsx -> CRM. Keep ONLY WhatsApp-available leads (Yes/likely),
# drop phone-only. Normalize phone to +965XXXXXXXX, auto-generate wa.me links.
# DELETE all old OSM leads (15) — replace entirely.
import openpyxl, re, json, os, glob, shutil

SRC = r"./Kuwait_Salon_WhatsApp_Leads.xlsx"
LEADS = os.environ.get("CLIENTS_DIR", r"D:\KB Rewaq Clients")
AGENCY = os.environ.get("AGENCY_DIR", r"D:\digitalfirst-agency")
BOSS = "96550703252"

wb = openpyxl.load_workbook(SRC)
ws = wb.active
kept = []
dropped = []
for r in range(2, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, 11)]
    name, area, phone, wa_link, wa_listed = vals[1], vals[2], vals[3], vals[4], vals[5]
    if not name:
        continue
    listed = str(wa_listed or "")
    if "Yes" not in listed and "likely" not in listed.lower():
        dropped.append(name)  # phone-only, skip
        continue
    # normalize phone: take first 8-digit chunk
    m = re.findall(r"\d{7,8}", str(phone))
    if not m:
        dropped.append(f"{name} (no-phone)")
        continue
    num8 = m[0]
    full = "965" + num8
    if len(num8) != 8:
        dropped.append(f"{name} (bad-phone:{phone})")
        continue
    wa = f"https://wa.me/{full}?text=" + "Hi%20" + re.sub(r'[^A-Za-z0-9]','%20',str(name)) + "%2C%20Jarvis%20here%20from%20KB%20Rewaq%20Digital.%20We%20build%20fresh%20websites%20%2B%20automation%20for%20salons%20in%20Kuwait.%20Want%20a%20free%20demo%3F"
    kept.append({
        "name_en": str(name).strip(),
        "area": str(area or "").strip(),
        "phone": full,
        "phone_disp": "+965****" + num8[-4:],
        "wa": wa,
        "ig": "", "site_url": "", "services": str(vals[7] or ""),
        "address": str(vals[8] or ""), "source": str(vals[9] or ""),
        "status": "lead_new",
    })

print(f"KEPT (WhatsApp): {len(kept)} | DROPPED (phone-only/bad): {len(dropped)}")
print("Dropped:", dropped)

# ---- DELETE all old OSM leads ----
old_dirs = glob.glob(os.path.join(LEADS, "*/"))
deleted = 0
for d in old_dirs:
    if os.path.basename(d.rstrip("/")) not in ("_scraped_phones.json", "_scraped_phones_backup.json"):
        shutil.rmtree(d, ignore_errors=True); deleted += 1
print(f"Deleted old OSM lead dirs: {deleted}")

# write new client folders (NESTED structure matching build_crm.py)
for k in kept:
    slug = re.sub(r"[^a-z0-9]+", "-", k["name_en"].lower()).strip("-")
    d = os.path.join(LEADS, slug)
    os.makedirs(d, exist_ok=True)
    rec = {
        "business": {"slug": slug, "name_en": k["name_en"], "name_ar": ""},
        "contact": {"phone": k["phone"], "phone_disp": k["phone_disp"], "wa": k["wa"]},
        "location": {"area_en": k["area"]},
        "online": {"site_url": "", "site_status": "live_fresh"},
        "pipeline": {"status": "lead_new", "source": "opencode_agent", "created": "2026-08-12"},
    }
    json.dump(rec, open(os.path.join(d, "client.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=2)

# build crm_leads.json
json.dump(kept, open(os.path.join(AGENCY, "crm_leads.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=2)
print(f"crm_leads.json written: {len(kept)} leads")

# save summary
json.dump({"kept": len(kept), "dropped": dropped}, open(os.path.join(AGENCY, "import_summary.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=2)
