#!/usr/bin/env python3
# Generate KB_Rewaq_Leads_v6.xlsx — "pahle jaisa" 8-column lead tracker.
# Reads all 94 leads from D:/KB Rewaq Clients/*/client.json + site-link.txt + dms/.
# 10 leads have live sites (site_status live*); 84 are pending_build (no real data yet).
# Phone cell -> wa.me link; Demo cell -> github site; DM cell -> ready message text.
import os, json, glob, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CRM = r"D:\KB Rewaq Clients"
OUT = r"D:\digitalfirst-agency\KB_Rewaq_Leads_v6.xlsx"
WA_YOU = "96550703252"
GH = "https://faizanbashar215.github.io/kb-rewaq-digital"

# ---- collect leads ----
leads = []
for d in sorted(glob.glob(os.path.join(CRM, "*/"))):
    cj = os.path.join(d, "client.json")
    if not os.path.isfile(cj):
        continue
    try:
        j = json.load(open(cj, encoding="utf-8"))
    except Exception:
        continue
    biz = j.get("business", {})
    con = j.get("contact", {})
    loc = j.get("location", {})
    onl = j.get("online", {})
    pipe = j.get("pipeline", {})
    slug = biz.get("slug") or os.path.basename(d.rstrip("/\\"))
    phone = con.get("phone", "") or ""
    phone_disp = con.get("phone_disp", "") or ""
    if not phone_disp and phone:
        phone_disp = ("+965 " + " ".join(re.findall(r"\d{1,4}", phone))) if phone.startswith("965") else phone
    # site url
    site = onl.get("site_url", "") or ""
    if not site:
        sl = os.path.join(d, "site-link.txt")
        if os.path.isfile(sl):
            site = open(sl, encoding="utf-8").read().strip()
    # live?
    status_raw = str(onl.get("site_status", "")).lower()
    has_data = isinstance(j.get("services"), list) and len(j.get("services", [])) > 0
    is_live = ("live" in status_raw) or (site and "github.io" in site and has_data)
    # DM draft
    dm_file = None
    dm_dir = os.path.join(d, "dms")
    if os.path.isdir(dm_dir):
        ds = sorted(glob.glob(os.path.join(dm_dir, "dm_*.txt")))
        if ds:
            dm_file = ds[-1]
    dm_text = open(dm_file, encoding="utf-8").read().strip() if (dm_file and is_live) else ""
    # For pending (no live data/site) do NOT invent a fake demo link — generic pitch only
    if not dm_text:
        if L_is_live_placeholder := (site and "github.io" in site and has_data):
            dm_text = (
                f"Hello {biz.get('name_en','')}! I built a free demo website for your salon.\n\n"
                f"Here is your demo: {site}\n\n"
                f"This is Faizan from KB Rewaq Digital. No pressure — if you like it we can take it further."
            )
        else:
            dm_text = (
                f"Hello {biz.get('name_en','')}! This is Faizan from KB Rewaq Digital — Kuwait.\n"
                f"We build websites + run Instagram/social + paid ads for ladies salons & spas.\n"
                f"Our packages start at just 35 KWD/month (Tier 1). 50% advance, 3-day delivery, cash/bank ok.\n"
                f"May I share the full packages? No pressure at all."
            )
    area = loc.get("area_en", "") or ""
    leads.append({
        "slug": slug,
        "name_en": biz.get("name_en", ""),
        "name_ar": biz.get("name_ar", ""),
        "area": area,
        "phone": phone,
        "phone_disp": phone_disp,
        "site": site,
        "is_live": bool(is_live and site),
        "dm": dm_text,
        "status": "Live" if (is_live and site) else "Pending Build",
    })

leads.sort(key=lambda x: (not x["is_live"], x["name_en"].lower()))

# ---- workbook ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "KB Rewaq Leads"
headers = ["#", "Business (EN)", "Business (AR)", "Area", "Phone (click→WhatsApp)",
           "Demo Website (click)", "Ready DM Message (link inside)", "Status"]
ws.append(headers)

green = Font(color="006100"); green_fill = PatternFill("solid", fgColor="C6EFCE")
blue = Font(color="0563C1", underline="single"); blue_fill = PatternFill("solid", fgColor="DDEBF7")
hdr_fill = PatternFill("solid", fgColor="1F4E78"); hdr_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")

for i, L in enumerate(leads, 1):
    from urllib.parse import quote
    phone_link = f"https://wa.me/{L['phone']}?text={quote(L['dm'])}" if L["phone"] else ""
    demo_link = L["site"] if L["is_live"] else ""
    row = [
        i,
        L["name_en"],
        L["name_ar"],
        L["area"],
        f'=HYPERLINK("{phone_link}","{L["phone_disp"] or "—"}")' if phone_link else (L["phone_disp"] or "—"),
        f'=HYPERLINK("{demo_link}","View Demo Site")' if demo_link else "Pending",
        L["dm"],
        L["status"],
    ]
    ws.append(row)
    r = ws.max_row
    # style
    for c in range(1, 9):
        ws.cell(r, c).border = border
        ws.cell(r, c).alignment = wrap
    if L["is_live"]:
        ws.cell(r, 5).font = green
        ws.cell(r, 5).fill = green_fill
        ws.cell(r, 6).font = blue
        ws.cell(r, 6).fill = blue_fill
        ws.cell(r, 8).font = Font(color="006100", bold=True)
    else:
        ws.cell(r, 8).font = Font(color="9C5700")

# header style
for c in range(1, 9):
    ws.cell(1, c).fill = hdr_fill
    ws.cell(1, c).font = hdr_font
    ws.cell(1, c).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

# column widths
widths = [5, 28, 22, 18, 22, 18, 60, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 30

# HOW TO USE sheet
htu = wb.create_sheet("HOW TO USE")
htu["A1"] = "KB REWAQ DIGITAL — LEAD TRACKER"
htu["A1"].font = Font(bold=True, size=14, color="1F4E78")
lines = [
    "",
    "1. Click the GREEN phone cell -> opens WhatsApp to that lead with your DM + their demo link pre-typed.",
    "2. Click the BLUE 'View Demo Site' cell -> opens the website you built for them.",
    "3. Send the DM on WhatsApp. The link is already inside the message.",
    "4. When they reply YES, visit their salon for a free overview, then discuss price & sign up.",
    "5. After sending, change Status to: replied / won / paid.",
    "",
    f"Your WA: +{WA_YOU}   |   Brand: KB Rewaq Digital",
    "",
    f"Total leads: {len(leads)}   |   Live sites: {sum(1 for x in leads if x['is_live'])}   |   Pending build: {sum(1 for x in leads if not x['is_live'])}",
]
for i, t in enumerate(lines, 2):
    htu.cell(i, 1).value = t
htu.column_dimensions["A"].width = 110

wb.save(OUT)
live = sum(1 for x in leads if x["is_live"])
print(f"SAVED {OUT}")
print(f"Total {len(leads)} | Live {live} | Pending {len(leads)-live}")
