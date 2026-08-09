#!/usr/bin/env python3
"""Build KB Rewaq lead Excel with click-to-WhatsApp + DM (link embedded) + site link.
DM rules (boss 2026-08-09): human/conversational tone, NO bullets/dashes/emoji-spam,
no upfront price, covers ALL 6 KB Rewaq services (Website, Post, Content, Marketing,
Ads, Strategy) naturally, closes with office-visit offer on YES.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import quote

GITHUB = "https://faizanbashar215.github.io/kb-rewaq-digital"

# Each lead: (en, ar, area, phone, disp, slug, opener)
# opener is a natural, personal WhatsApp opener (no emoji/dash/bullet).
LEADS = [
    ("KAHI Ladies Beauty Salon", "صالون كاهي", "Salmiya Block 2", "96550770458", "+965 5077 0458", "kahi",
     "Hi, this is Faizan. Heard good things about KAHI and wanted to show you something I made."),
    ("Mahima Ladies Salon", "صالون ماهيما", "Salmiya Block 10", "96551413855", "+965 5141 3855", "mahima",
     "Hi, Faizan here. Your spa packages look amazing, your clients would love seeing them online."),
    ("Paulita Spa Beauty", "بوليتا سبا", "Salmiya Block 9", "96569649660", "+965 6964 9660", "paulita",
     "Hello, this is Faizan. Your salon has such a relaxing vibe, it should have a beautiful site too."),
    ("Neweves Salon", "صالون نيو إيفز", "Hawally Tunis St", "96594147140", "+965 9414 7140", "neweves",
     "Hi, Faizan here. You have got great prices and work, just thought a site would bring more people in."),
    ("Yours Salon", "صالون يورز", "Maidan Hawally", "96597200323", "+965 9720 0323", "yours",
     "Hi, this is Faizan. Your regulars clearly love you, a website would just help new ones find you."),
]

# Human, conversational DM body covering ALL 6 KB Rewaq services (no bullets/dashes/emoji-spam)
BODY = (
    "\n\nThis is Faizan, I run a small digital studio here in Kuwait called KB Rewaq. "
    "I made a little demo website for your salon just to show you what it could look like, "
    "totally free and no strings attached.\n\n"
    "If you like it, I can handle the whole thing for you. A proper website with your name, "
    "your services and prices and online booking in both Arabic and English. I also look after "
    "your social posts and reels, you just send me photos from the salon and I take care of the "
    "content so you stay active and keep pulling in new clients. On top of that I do the video "
    "work, the marketing, the paid ads and the strategy, so it is not just about posting, it is "
    "about actually growing your salon.\n\n"
    "No pressure at all, but if you are curious I can come by the salon for 15 minutes, show you "
    "the plan properly and we take it from there.\n\n"
    "Just reply with a yes and I will get it started :)"
)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "KB Rewaq Leads"

hdr_fill = PatternFill("solid", fgColor="6A0DAD")
hdr_font = Font(bold=True, color="FFFFFF", size=12)
cell_font = Font(size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wa_fill = PatternFill("solid", fgColor="25D366")
site_fill = PatternFill("solid", fgColor="1E90FF")

headers = ["#", "Business (EN)", "Business (AR)", "Area", "Phone (click→WhatsApp)",
           "Demo Website (click)", "Ready DM Message (link inside)", "Status"]
widths = [4, 26, 18, 16, 24, 30, 60, 14]

for c, h in enumerate(headers, 1):
    cell = ws.cell(1, c, h)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center; cell.border = border
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

for i, (en, ar, area, phone, disp, slug, opener) in enumerate(LEADS, 1):
    r = i + 1
    site_url = f"{GITHUB}/{slug}/"
    dm = opener + BODY + f"\n\nHere is your free demo: {site_url}"
    wa_url = f"https://wa.me/{phone}?text={quote(dm)}"
    vals = [i, en, ar, area, disp, "View Demo Site", dm, "Lead"]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v)
        cell.font = cell_font; cell.border = border
        cell.alignment = left if c in (2, 3, 4, 7) else center
    ws.cell(r, 5).hyperlink = wa_url
    ws.cell(r, 5).fill = wa_fill
    ws.cell(r, 5).font = Font(size=11, bold=True, color="FFFFFF")
    ws.cell(r, 6).hyperlink = site_url
    ws.cell(r, 6).fill = site_fill
    ws.cell(r, 6).font = Font(size=11, bold=True, color="FFFFFF")
    ws.cell(r, 6).alignment = center
    ws.row_dimensions[r].height = 170

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{len(LEADS)+1}"

ws2 = wb.create_sheet("HOW TO USE")
instructions = [
    ("KB REWAQ DIGITAL — LEAD TRACKER", True),
    ("", False),
    ("1. Click the GREEN phone cell -> opens WhatsApp to that lead with your DM + their demo link pre-typed.", False),
    ("2. Click the BLUE 'View Demo Site' cell -> opens the 3D website you built for them.", False),
    ("3. Send the DM on WhatsApp. The link is already inside the message.", False),
    ("4. When they reply YES, visit their salon for a free overview, then discuss price & sign up.", False),
    ("5. After sending, change Status to: replied / won / paid.", False),
    ("", False),
    ("Your WA: +965 50703252   |   Brand: KB Rewaq Digital", False),
]
for i, (txt, bold) in enumerate(instructions, 1):
    c = ws2.cell(i, 1, txt)
    c.font = Font(bold=bold, size=14 if bold else 11)
    ws2.column_dimensions["A"].width = 100

wb.save("KB_Rewaq_Leads_v6.xlsx")
print("✅ KB_Rewaq_Leads_v6.xlsx created with", len(LEADS), "leads (v10 advanced sites, all 6 KB Rewaq services in DM)")
